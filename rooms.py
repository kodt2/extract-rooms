import openpyxl
from openpyxl.cell.cell import MergedCell
import datetime
import fpdf


def numb_to_month(n):
    months = {
        1: "января", 2: "февраля", 3: "марта", 4: "апреля",
        5: "мая", 6: "июня", 7: "июля", 8: "августа",
        9: "сентября", 10: "октября", 11: "ноября", 12: "декабря"
    }
    n = int(n)
    return months.get(n, "NaM")
def month_to_numb(str):
    months = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4,
    "мая": 5, "июня": 6, "июля": 7, "августа": 8,
    "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12
    }
    return months.get(str, "NaM")
def time_to_pairs(start_time, end_time):
    def time_to_float(t):
        hours, minutes = map(int, t.split(":"))
        if int(hours) > 24:
            raise Exception("hours should be less or equal than 24")
        if int(minutes) > 60:
            raise Exception("minutes should be less or equal than 60")
        return hours + minutes / 60

    fl_start_time = time_to_float(start_time)
    fl_end_time = time_to_float(end_time)
    #print(fl_start_time, fl_end_time)
    if (fl_start_time < 7.0 or fl_end_time < 7.0):
        raise Exception("Лобач открывается в 7:30")
    if (fl_start_time > 22 or fl_end_time > 22):
        raise Exception("Лобач закрывается в 21")
    # Время начала и конца каждой пары
    pairs = [(7.5, 9+10/60), (9.0, 10+50/60), (10.0 + 4 / 6, 13.0), (13.0, 14.0 + 40 / 60),
             (14.0 + 30 / 60, 16.0 + 20 / 60), (16.0 + 10 / 60, 18.0), (17.0+5/6, 19.0 + 40 / 60),
             (19.0 + 40 / 60, 21+20/60)]
    rev_pairs = [(19.0 + 40 / 60, 21 + 20 / 60),
             (18.0, 19.0 + 40 / 60),
             (16.0 + 10 / 60, 18.0),
             (14.0 + 40 / 60, 16.0 + 20 / 60),
             (13.0, 14.0 + 30 / 60),
             (10.0 + 4 / 6, 13.0),
             (9.0, 10 + 50 / 60),
             (7.5, 9 + 10 / 60)]

    print(pairs)
    print(0)
    for i in range(7, -1,-1):
        print(pairs[i])
        if pairs[i][0]<=fl_start_time<=pairs[i][1]:
            num_st_p = i+1
            break
        else :
            num_st_p = 1
    print(num_st_p)
    for i in range(0, 8):
        if pairs[i][0] <= fl_end_time <= pairs[i][1]:
            num_en_p = i + 1
            break
        else :
            num_en_p = 8
    print(num_en_p)
    print(start_time, end_time, list(range(num_st_p, num_en_p + 1)))
    return list(range(num_st_p, num_en_p + 1))


def get_aud( vihod ):
    # Используем set comprehension для создания множества и set -> sorted list
    inp=[]
    for i in vihod:
        inp.append([i[1], i[2], i[3], i[4], i[5], i[6], i[7], i[8]])
    s = sorted({int(i[7]) for i in inp})

    # Объединяем элементы списка в строку через запятую
    korp = ", ".join(map(str, s))

    return korp, inp


def nedeed_text1(pdf, korp, KrotovaNV, KrotovNV, Zolotuh_name, VolkovuAE, Borisov_name, predsed_name, predsed_phone):
    pdf.cell(200, 13, txt="", ln=1, align="R")
    pdf.cell(190, 6, txt="Начальнику управления безопасности", ln=1, align="R")
    pdf.cell(190, 6, txt="ННГУ им. Н.И. Лобачевского", ln=1, align="R")
    pdf.cell(190, 6, txt=VolkovuAE, ln=1, align="R")
    pdf.cell(190, 6, txt="от заместителя директора по учебно-воспитательной работе", ln=1, align="R")
    pdf.cell(190, 6, txt="Института Информационных", ln=1, align="R")
    pdf.cell(190, 6, txt="Технологий, Математики и Механики", ln=1, align="R")
    pdf.cell(190, 6, txt=KrotovaNV, ln=1, align="R")
    pdf.cell(190, 8, txt="", ln=1, align="C")
    pdf.cell(190, 15, txt="Служебная записка.", ln=1, align="C")
    pdf.cell(10)
    pdf.multi_cell(180, 6, txt="            Институт Информационных технологий математики"
                               " и механики просит\n "
                               "Вас предоставить доступ в "
                               "аудитории корпуса № " + korp + " ННГУ им. Н.И. Лобачевского "
                                                               "для проведения собрания Студенческого Совета ИИТММ.",
                   align="L")
    pdf.cell(190, 6, txt="", ln=1, align="l")
    return pdf


def nedeed_text2(pdf, KrotovaNV, KrotovNV, Zolotuh_name, VolkovuAE, Borisov_name, predsed_name, predsed_phone):
    pdf.cell(10)
    pdf.cell(190, 6, txt="Ответственные:", ln=1, align="l")
    pdf.cell(10)
    pdf.cell(190, 6, txt=Zolotuh_name + ", " + Borisov_name + ", " + KrotovNV, ln=1, align="l")
    pdf.cell(190, 6, txt="", ln=1, align="l")
    pdf.cell(10)
    pdf.cell(100, 6, txt="Заместитель директора ИИТММ", ln=1, align="l")
    pdf.cell(10)
    pdf.cell(100, 6, txt="по учебно-воспитательной работе", align="l")
    pdf.cell(40)
    pdf.cell(90, 6, txt=KrotovNV, ln=1, align="r")
    pdf.cell(190, 6, txt="", ln=1, align="l")
    pdf.cell(190, 6, txt="Председатель СС ИИТММ: " + predsed_name + " " + predsed_phone, ln=1, align="R")
    return pdf


def cycle_out_rooms(pdf, arr, start, end):
    for i in range(start, end):
        pdf.cell(10)
        pdf.cell(190, 6, txt=arr[i][0] + " " + arr[i][1] + " " + str(arr[i][2])+' '+ arr[i][3] + u" c " + arr[i][4] + u" до " + arr[i][5]
                             + u" ауд. " + arr[i][6] + u" корпуса № " + str(arr[i][7]), ln=1, align="l")
    pdf.cell(190, 6 * (17 - end + start), txt="", ln=1, align="C")
    return pdf


def work_rooms(arr, a, c, b, d, e, predsed_name, predsed_phone):
    year = datetime.date.today().year
    print('-1')
    try:
        work6 = openpyxl.load_workbook("inp6.xlsx")
    except Exception as e:
        print(e)
    print('-2')
    work2 = openpyxl.load_workbook("inp2.xlsx")
    print('-2')
    sheet2 = work2[u'Загруженность аудиторий']
    print('-3')
    sheet6 = work6[u'Загруженность аудиторий']
    print('0')
    week2 = {}
    old_coord = 2
    val = sheet2['A2'].value.split(' ')
    for cell in sheet2['A']:
        if cell.coordinate == 'A1':
            continue
        if cell.coordinate == 'A2':
            continue
        # print(cell.coordinate, isinstance(cell, MergedCell), cell.value)
        if cell.value is not None:
            # print(val, cell.coordinate)
            new_coord = int(cell.coordinate[1:])
            key = (val[0], val[1])
            week2[key] = [val[4], old_coord, new_coord - 1]
            old_coord = int(cell.coordinate[1:])
            val = cell.value.split(' ')
        if not isinstance(cell, MergedCell) and cell.value is None:
            break
    '''for key, value in week2.items():
        print(f"Key2: {key}, Value2: {value}")'''
    print('8')
    week6 = {}
    old_coord = 2
    val = sheet6['A2'].value.split(' ')
    for cell in sheet6['A']:
        if (cell.coordinate == ('A1')):
            continue
        if (cell.coordinate == ('A2')):
            continue
        # print(cell.coordinate, isinstance(cell, MergedCell), cell.value)
        if cell.value != None:
            # print(val, cell.coordinate)
            new_coord = int(cell.coordinate[1:])
            key = (val[0], val[1])
            week6[key] = [ val[4], old_coord, new_coord - 1]
            old_coord = int(cell.coordinate[1:])
            val = cell.value.split(' ')
        if not isinstance(cell, MergedCell) and cell.value == None:
            break
    '''for key, value in week6.items():
        print(f"Key6: {key}, Value6: {value}")'''

    coords_aud2=[]
    valid_aud_values2_L = ['105', '307', '309', '314', '324', '328']
    valid_aud_values2_S = [ '317', '318', '322']

    for cell in sheet2['1']:
        if (cell.value in valid_aud_values2_L):
            coords_aud2.append((cell.value,cell.coordinate[:-1],2 ,'big'))
        if (cell.value in valid_aud_values2_S):
            coords_aud2.append((cell.value, cell.coordinate[:-1],2 , 'small'))
        if cell.value == None:
            break
    #print(coords_aud2)
    print('8')
    coords_aud6=[]
    valid_aud_values6_L = ['502','513']
    valid_aud_values6_S = ['506', '508', '509',
                        '511', '512', '514', '515']
    for cell in sheet6['1']:
        if (cell.value in valid_aud_values6_L):
            coords_aud6.append((cell.value,cell.coordinate[:-1],6 ,'big'))
        if (cell.value in valid_aud_values6_S):
            coords_aud6.append((cell.value,cell.coordinate[:-1],6 ,'small'))
        if cell.value == None:
            break
    print('8')
    '''print(coords_aud6)'''

    for key in week2.keys():
        aud=[]
        first_pair = int(sheet2['B' + str(week2[key][1])].value[0])
        for coord in coords_aud2:
            data=[]
            if first_pair != 1:
                for i in range(0, first_pair - 1):
                    data.append('')
            for cell in sheet2[coord[1] + str(week2[key][1]):coord[1] + str(week2[key][2])]:
                data.append(cell[0].value)
            while len(data) <= 7:
                data.append('')
            aud.append((coord[0], coord[2], coord[3], data))
        week2[key] = (week2[key][0],aud)
    '''for key, value in week2.items():
        print(f"Key2: {key}, Value2: {value}")'''
    print('9')
    for key in week6.keys():
        aud=[]
        first_pair = int(sheet6['B'+str(week6[key][1])].value[0])
        for coord in coords_aud6:
            #print(first_pair, coord[0], key)
            data=[]
            if first_pair!=1:
                for i in range(0,first_pair-1):
                    data.append('')
            for cell in sheet6[coord[1] + str(week6[key][1]):coord[1] + str(week6[key][2])]:
                data.append(cell[0].value)
            while len(data) <= 7:
                data.append('')
            aud.append((coord[0], coord[2], coord[3], data))
        week6[key] = (week6[key][0],aud)
    '''for key, value in week6.items():
        print(f"Key6: {key}, Value6: {value}")'''

    req = []
    print('10')
    for i in arr:
        try:
            month = numb_to_month((i[3].split('.'))[1])
        except Exception("wrong date format Expected: [date.month]"):
            print("wrong date format Expected: [date.month]")
        try:
            req.append(
                [i[0] + " " + i[1], i[2], str(int(i[3].split('.')[0])), month, year, time_to_pairs(i[4], i[5]), i[6],
                 i[4], i[5]])
        except Exception("wrong time format Expected [hours:minutes]"):
            print("wrong time format Expected [hours:minutes]")

    vihod = []

    print(req)
    def typical():
        all_free = False
        find_in_2 = False
        for aud in data2[1]:
            if aud[0] == r[6]:
                all_free = all(aud[3][i - 1] == '' for i in r[5])
            if all_free:
                find_in_2 = True
                vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')',  r[7], r[8], aud[0],
                              aud[1]])
                # print(aud[0], aud[1], aud[2])
                for i in r[5]:
                    aud[3][i-1] = 'CCC'
                break
        if not (find_in_2):
            vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])

    for r in req:
        if not (r[2], r[3]) in week2.keys():
            vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + ')', r[7], r[8], "", "day_not_in_schedule"])
            continue
        if not (r[2], r[3]) in week6.keys():
            vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + ')', r[7], r[8], "", "day_not_in_schedule"])
            continue
        data2 = week2[(r[2],r[3])]
        data6 = week6[(r[2], r[3])]
        #print(data2)
        day_of_week = data2[0]
        match r[6]:
            case 'big':
                all_free = False
                find_in_6 = False
                find_in_2 = False
                for aud in data6[1]:
                    if aud[2] == 'big':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_6 = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        #print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        break
                if not find_in_6:
                    for aud in data2[1]:
                        if aud[2] == 'big':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_2 = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            #print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            break
                if not (find_in_6) and not (find_in_2):
                    vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            case 'big2':
                all_free = False
                find_in_2 = False
                for aud in data2[1]:
                    if aud[2] == 'big':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_2 = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        # print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        break
                if  not (find_in_2):
                    vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            case '105':
                typical()
            case '307':

                typical()
            # some code
            case '309':

                typical()
            # some code
            case '314':

                typical()
            # some code
            case '317':

                typical()
            # some code
            case '318':

                typical()
            # some code
            case '322':

                typical()
            case '324':

                typical()
            # some code
            case '328':

                typical()
            # some code
            case 'any':
                all_free = False
                find_in_6 = False
                find_in_2 = False
                find_in_small = False
                for aud in data6[1]:
                    if aud[2] == 'small':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_6 = True
                        find_in_small = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        #print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        break
                if not find_in_6:
                    for aud in data2[1]:
                        if aud[2] == 'small':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_small = True
                            find_in_2 = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            #print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            break
                if not find_in_small:
                    for aud in data6[1]:
                        if aud[2] == 'big':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_6 = True
                            find_in_small = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            #print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            break
                    if not find_in_6:
                        for aud in data2[1]:
                            if aud[2] == 'big':
                                all_free = all(aud[3][i - 1] == '' for i in r[5])
                            if all_free:
                                find_in_small = True
                                find_in_2 = True
                                vihod.append(
                                    [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                     aud[1]])
                                #print(aud[0], aud[1], aud[2])
                                for i in r[5]:
                                    aud[3][i - 1] = 'CCC'
                                break
                if not (find_in_6) and not (find_in_2):
                    vihod.append(
                        [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            # some code
            case 'any2':
                all_free = False
                find_in_2 = False
                find_in_small = False
                for aud in data2[1]:
                    if aud[2] == 'small':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_small = True
                        find_in_2 = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        # print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        break
                if not find_in_small:
                    for aud in data2[1]:
                        if aud[2] == 'big':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_2 = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            # print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            break
                if not (find_in_2):
                    vihod.append(
                        [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            # some code
            case _:
                vihod.append(
                [r[0] + " " + r[1], int(r[2]), r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "wrong postfix " + "'" + r[6] + "'"])

    #print(vihod)
    return (vihod)

def work_pdf(arr, KaNV, KNV, Z_name, VAE, B_name, predsed_name, predsed_phone):
    year = datetime.date.today().year
    work6 = openpyxl.load_workbook("inp6.xlsx")
    work2 = openpyxl.load_workbook("inp2.xlsx")
    sheet2 = work2[u'Загруженность аудиторий']
    sheet6 = work6[u'Загруженность аудиторий']

    week2 = {}
    old_coord = 2
    val = sheet2['A2'].value.split(' ')
    for cell in sheet2['A']:
        if cell.coordinate == 'A1':
            continue
        if cell.coordinate == 'A2':
            continue
        # print(cell.coordinate, isinstance(cell, MergedCell), cell.value)
        if cell.value is not None:
            # print(val, cell.coordinate)
            new_coord = int(cell.coordinate[1:])
            key = (val[0], val[1])
            week2[key] = [val[4], old_coord, new_coord - 1]
            old_coord = int(cell.coordinate[1:])
            val = cell.value.split(' ')
        if not isinstance(cell, MergedCell) and cell.value is None:
            break
    '''for key, value in week2.items():
        print(f"Key2: {key}, Value2: {value}")'''

    week6 = {}
    old_coord = 2
    val = sheet6['A2'].value.split(' ')
    for cell in sheet6['A']:
        if (cell.coordinate == ('A1')):
            continue
        if (cell.coordinate == ('A2')):
            continue
        # print(cell.coordinate, isinstance(cell, MergedCell), cell.value)
        if cell.value != None:
            # print(val, cell.coordinate)
            new_coord = int(cell.coordinate[1:])
            key = (val[0], val[1])
            week6[key] = [val[4], old_coord, new_coord - 1]
            old_coord = int(cell.coordinate[1:])
            val = cell.value.split(' ')
        if not isinstance(cell, MergedCell) and cell.value == None:
            break
    '''for key, value in week6.items():
        print(f"Key6: {key}, Value6: {value}")'''

    coords_aud2 = []
    valid_aud_values2_L = ['105', '307', '309', '314', '324', '328']
    valid_aud_values2_S = ['317', '318', '322']

    for cell in sheet2['1']:
        if (cell.value in valid_aud_values2_L):
            coords_aud2.append((cell.value, cell.coordinate[:-1], 2, 'big'))
        if (cell.value in valid_aud_values2_S):
            coords_aud2.append((cell.value, cell.coordinate[:-1], 2, 'small'))
        if cell.value == None:
            break
    # print(coords_aud2)

    coords_aud6 = []
    valid_aud_values6_L = ['502', '513']
    valid_aud_values6_S = ['506', '508', '509',
                           '511', '512', '514', '515']
    for cell in sheet6['1']:
        if (cell.value in valid_aud_values6_L):
            coords_aud6.append((cell.value, cell.coordinate[:-1], 6, 'big'))
        if (cell.value in valid_aud_values6_S):
            coords_aud6.append((cell.value, cell.coordinate[:-1], 6, 'small'))
        if cell.value == None:
            break
    '''print(coords_aud6)'''

    for key in week2.keys():
        aud = []
        first_pair = int(sheet2['B' + str(week2[key][1])].value[0])
        for coord in coords_aud2:
            data = []
            if first_pair!=1:
                for i in range(0,first_pair-1):
                    data.append('')
            for cell in sheet2[coord[1] + str(week2[key][1]):coord[1] + str(week2[key][2])]:
                data.append(cell[0].value)
            while len(data) <= 7:
                data.append('')
            aud.append((coord[0], coord[2], coord[3], data))
        week2[key] = (week2[key][0], aud)
    '''for key, value in week2.items():
        print(f"Key2: {key}, Value2: {value}")'''
    for key in week6.keys():
        aud = []
        first_pair = int(sheet6['B' + str(week6[key][1])].value[0])
        for coord in coords_aud6:
            data = []
            if first_pair!=1:
                for i in range(0,first_pair-1):
                    data.append('')
            for cell in sheet6[coord[1] + str(week6[key][1]):coord[1] + str(week6[key][2])]:
                data.append(cell[0].value)
            while len(data) <= 7:
                data.append('')
            aud.append((coord[0], coord[2], coord[3], data))
        week6[key] = (week6[key][0], aud)
    '''for key, value in week6.items():
        print(f"Key6: {key}, Value6: {value}")'''

    req = []
    for i in arr:
        #print(i)
        try:
            month = numb_to_month((i[3].split('.'))[1])
        except Exception("wrong date format Expected: [date.month]"):
            print("wrong date format Expected: [date.month]")
        try:
            req.append(
                [i[0] + " " + i[1], i[2], str(int(i[3].split('.')[0])), month, year, time_to_pairs(i[4], i[5]), i[6],
                 i[4], i[5]])
        except Exception("wrong time format Expected hours:minutes"):
            print("wrong time format Expected")

    vihod = []

    # print(req)
    def typical():
        all_free = False
        find_in_2 = False
        for aud in data2[1]:
            if aud[0] == r[6]:
                all_free = all(aud[3][i - 1] == '' for i in r[5])
            if all_free:
                find_in_2 = True
                vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                              aud[1]])
                # print(aud[0], aud[1], aud[2])
                for i in r[5]:
                    aud[3][i-1] = 'CCC'
                break
        if not (find_in_2):
            vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])

    for r in req:
        if not (r[2], r[3]) in week2.keys():
            vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + ')', r[7], r[8], "", "day_not_in_schedule"])
            continue
        if not (r[2], r[3]) in week6.keys():
            vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + ')', r[7], r[8], "", "day_not_in_schedule"])
            continue
        data2 = week2[(r[2], r[3])]
        data6 = week6[(r[2], r[3])]
        # print(data2)
        day_of_week = data2[0]
        match r[6]:
            case 'big':
                all_free = False
                find_in_6 = False
                find_in_2 = False
                for aud in data6[1]:
                    if aud[2] == 'big':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_6 = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        # print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        #print(aud)
                        break
                if not find_in_6:
                    for aud in data2[1]:
                        if aud[2] == 'big':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_2 = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            # print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            break
                if not (find_in_6) and not (find_in_2):
                    vihod.append(
                        [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            case 'big2':
                all_free = False
                find_in_2 = False
                for aud in data2[1]:
                    if aud[2] == 'big':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_2 = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        # print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        break
                if not (find_in_2):
                    vihod.append(
                        [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            case '105':
                typical()
            case '307':

                typical()
            # some code
            case '309':

                typical()
            # some code
            case '314':

                typical()
            # some code
            case '317':

                typical()
            # some code
            case '318':

                typical()
            # some code
            case '322':

                typical()
            case '324':

                typical()
            # some code
            case '328':

                typical()
            # some code
            case 'any':
                all_free = False
                find_in_6 = False
                find_in_2 = False
                find_in_small = False
                for aud in data6[1]:
                    if aud[2] == 'small':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_6 = True
                        find_in_small = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        # print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        #print(aud)
                        break
                if not find_in_6:
                    for aud in data2[1]:
                        if aud[2] == 'small':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_small = True
                            find_in_2 = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            # print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            #print(aud)
                            break
                if not find_in_small:
                    for aud in data6[1]:
                        if aud[2] == 'big':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_6 = True
                            find_in_small = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            # print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            #print(aud)
                            break
                    if not find_in_6:
                        for aud in data2[1]:
                            if aud[2] == 'big':
                                all_free = all(aud[3][i - 1] == '' for i in r[5])
                            if all_free:
                                find_in_small = True
                                find_in_2 = True
                                vihod.append(
                                    [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                     aud[1]])
                                # print(aud[0], aud[1], aud[2])
                                for i in r[5]:
                                    aud[3][i - 1] = 'CCC'
                                #print(aud)
                                break
                if not (find_in_6) and not (find_in_2):
                    vihod.append(
                        [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            # some code
            case 'any2':
                all_free = False
                find_in_2 = False
                find_in_small = False
                for aud in data2[1]:
                    if aud[2] == 'small':
                        all_free = all(aud[3][i - 1] == '' for i in r[5])
                    if all_free:
                        find_in_small = True
                        find_in_2 = True
                        vihod.append([r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                      aud[1]])
                        # print(aud[0], aud[1], aud[2])
                        for i in r[5]:
                            aud[3][i - 1] = 'CCC'
                        break
                if not find_in_small:
                    for aud in data2[1]:
                        if aud[2] == 'big':
                            all_free = all(aud[3][i - 1] == '' for i in r[5])
                        if all_free:
                            find_in_2 = True
                            vihod.append(
                                [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], aud[0],
                                 aud[1]])
                            # print(aud[0], aud[1], aud[2])
                            for i in r[5]:
                                aud[3][i - 1] = 'CCC'
                            break
                if not (find_in_2):
                    vihod.append(
                        [r[0] + " " + r[1], r[2], r[3], year, '(' + day_of_week + ')', r[7], r[8], "", "no free room"])
            # some code
            case _:
                vihod.append(
                    [r[0] + " " + r[1], int(r[2]), r[3], year, '(' + day_of_week + ')', r[7], r[8], "",
                     "wrong postfix " + "'" + r[6] + "'"])

    #print(vihod)
    vihod.sort(key=lambda x: x[1])
    #print(vihod)
    vihod.sort(key=lambda x: month_to_numb(x[2]))
    #print(vihod)
    arr = list()
    korp = ''
    try:
        korp, arr = get_aud(vihod)
    except Exception:
        print("no such room")
        raise Exception("no such room")
    pdf = fpdf.FPDF(format='letter')
    pdf.add_page()
    pdf.add_font('Times', '', 'times.ttf', uni=True)
    pdf.set_font("Times", size=14)
    pdf = nedeed_text1(pdf, korp, KaNV, KNV, Z_name, VAE, B_name, predsed_name,
                       predsed_phone)

    if len(arr) <= 15:
        pdf = cycle_out_rooms(pdf, arr, 0, len(arr))
        pdf = nedeed_text2(pdf, KaNV, KNV, Z_name, VAE, B_name, predsed_name, predsed_phone)
    else:
        count_pages = len(arr) // 15
        if (len(arr) % 14 != 0):
            count_pages += 1
        pdf = cycle_out_rooms(pdf, arr, 0, 15)
        pdf = nedeed_text2(pdf, KaNV, KNV, Z_name, VAE, B_name, predsed_name, predsed_phone)

        for i in range(1, count_pages - 1):
            pdf.add_page()
            pdf = nedeed_text1(pdf, korp, KaNV, KNV, Z_name, VAE, B_name, predsed_name,
                               predsed_phone)
            pdf = cycle_out_rooms(pdf, arr, i * 15, i * 15 + 15)
            pdf = nedeed_text2(pdf, KaNV, KNV, Z_name, VAE, B_name, predsed_name,
                               predsed_phone)
        pdf.add_page()
        pdf = nedeed_text1(pdf, korp, KaNV, KNV, Z_name, VAE, B_name, predsed_name,
                           predsed_phone)
        pdf = cycle_out_rooms(pdf, arr, (count_pages - 1) * 15, len(arr))
        pdf = nedeed_text2(pdf, KaNV, KNV, Z_name, VAE, B_name, predsed_name, predsed_phone)

    name = "pdf-s\Служебная записка "+datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S") +".pdf"
    pdf.output(name)
    print('создание pdf')
    return vihod, name

