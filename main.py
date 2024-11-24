
try:
    import openpyxl
except:
    print ("use 'pip install openpyxl'")
    quit()
try:
    import datetime
except:
    print("use 'pip install datetime'")
    quit()
try:
    import fpdf
except:
    print("use 'pip install fpdf'")
    quit()


def numb_to_month(n):
    n = int(n)
    if (n == '1' or n == 1):
        return u"января"
    if (n == '2' or n == 2):
        return u"февраля"
    if (n == '3' or n == 3):
        return u"марта"
    if (n == '4' or n == 4):
        return u"апреля"
    if (n == '5' or n == 5):
        return u"мая"
    if (n == '6' or n == 6):
        return u"июня"
    if (n == '7' or n == 7):
        return u"июля"
    if (n == '8' or n == 8):
        return u"августа"
    if (n == '9' or n == 9):
        return u"сентября"
    if (n == '10' or n == 10):
        return u"октября"
    if (n == '11' or n == 11):
        return u"ноября"
    if (n == '12' or n == 12):
        return u"декабря"
    return "NaM"


def time_to_pairs(start_time, end_time):
    fl_start_time = float(start_time.split(":")[0]) + float(start_time.split(":")[1]) / 60
    fl_end_time = float(end_time.split(":")[0]) + float(end_time.split(":")[1]) / 60
    s1, e1 = 7.5, 7.5 + 1.5
    s2, e2 = e1 + 1 / 6, e1 + 1 / 6 + 1.5
    s3, e3 = e2 + 1 / 6, e2 + 1 / 6 + 1.5
    s4, e4 = e3 + 4 / 6, e3 + 4 / 6 + 1.5
    s5, e5 = e4 + 1 / 6, e4 + 1 / 6 + 1.5
    s6, e6 = e5 + 1 / 6, e5 + 1 / 6 + 1.5
    s7, e7 = e6 + 1 / 6, e6 + 1 / 6 + 1.5
    s8, e8 = e7 + 1 / 6, e7 + 1 / 6 + 1.5
    st_p = [s1, s2, s3, s4, s5, s6, s7, s8]
    en_p = [e1, e2, e3, e4, e5, e6, e7, e8]
    num_st_p = 1
    num_en_p = 8
    for i in range(0, len(st_p)):
        if st_p[i] <= fl_start_time:
            num_st_p = i + 1
    for i in range(0, len(en_p)):
        if en_p[i] > fl_end_time:
            num_en_p = i + 1
        if en_p[i] == fl_end_time:
            num_en_p = i + 1
    res = []
    for i in range(num_st_p, num_en_p + 1):
        res.append(int(i))
    return res


year = datetime.datetime.today().year
work2 = openpyxl.load_workbook("inp6.xlsx")
work6 = openpyxl.load_workbook("inp2.xlsx")
sheet2 = work2[u'Загруженность аудиторий']
sheet6 = work6[u'Загруженность аудиторий']
week = []
week6 = []
val = ''
old_coord = int("A1"[1:])
for cellObj in sheet2['A2':'A100']:
    for cell in cellObj:
        if cell.value != sheet2['A3'].value:
            val = cell.value.split(' ')
            new_coord = int(cell.coordinate[1:])
            week.append([val[0], val[1], val[2], val[4], old_coord, new_coord])
            old_coord = int(cell.coordinate[1:])
#print(week)
for i in range(len(week) - 1):
    week[i][4] = week[i + 1][4]
    week[i][5] = week[i + 1][5] - 1
#print(week)
week[len(week) - 1][4] = week[len(week) - 1][5]
week[len(week) - 1][5] = week[len(week) - 1][5]+7
#print(week)
val = ''
old_coord = int("A1"[1:])
for cellObj in sheet6['A2':'A100']:
    for cell in cellObj:
        if cell.value != sheet6['A3'].value:
            val = cell.value.split(' ')
            new_coord = int(cell.coordinate[1:])
            week6.append([val[0], val[1], val[2], val[4], old_coord, new_coord])
            old_coord = int(cell.coordinate[1:])
for i in range(len(week6) - 1):
    week6[i][4] = week6[i + 1][4]
    week6[i][5] = week6[i + 1][5] - 1
week6[len(week6) - 1][4] = week6[len(week6) - 1][5]
week6[len(week6) - 1][5] = week6[len(week6) - 1][5]+7
for i in week:
    room_data = []
    for cellObj in sheet2['W' + str(i[4]):'AE' + str(i[5])]:
        help = []
        for cell in cellObj:
            if (cell.value == None):
                help.append('')
            else:
                help.append(cell.value)
        room_data.append(help)
    i.append(room_data)
    while len(i[6]) <= 7:
        i[6].append(['', '', '', '', '', '', '', '', ''])
#print(week6)
for i in week6:
    room_data = []
    for cellObj in sheet6['C' + str(i[4]):'AT' + str(i[5])]:
        help = []
        for cell in cellObj:
            if (cell.value == None):
                help.append('')
            else:
                help.append(cell.value)
        room_data.append(help)
    i.append(room_data)
    while len(i[6]) <= 7:
        i[6].append(
            ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '',
             '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''])
#print("###########################################")
#print(week6)
with open('needed rooms.txt', encoding='utf-8') as f:
    arr = f.read()
    arr = arr.split('\n')
    for i in range(len(arr)):
        arr[i] = arr[i].split(' ')

req = []
for i in arr:
    month = numb_to_month((i[3].split('.'))[1])
    req.append([i[0], i[1], str(int(i[3].split('.')[0])), month, year, time_to_pairs(i[4], i[5]), i[6], i[4], i[5]])
vihod = []
for i in req:
    if i[6] == 'big':
        day = i[2]
        month = i[3]
        year = i[4]
        pairs = i[5]
        also_find = 0
        right_week = 0
        for j in week:
            if (day == j[0] and month == j[1] and year == int(j[2])):
                cvobodn = 1
                right_week = 1
                for p in i[5]:
                    if (j[6][p - 1][6] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][6] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 513, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][0] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][0] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 502, 6])
                    also_find = 1
                    break
        if (also_find == 0):
            for j in week6:
                if (day == j[0] and month == j[1] and year == int(j[2])):
                    cvobodn = 1
                    right_week = 1
                    for p in i[5]:
                        if (j[6][p - 1][0] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][0] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 105, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][36] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][36] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 314, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][43] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][43] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 328, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][32] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][32] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 307, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][42] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][42] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 324, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][33] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][33] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 309, 2])
                        break
                    if cvobodn == 0:
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no free room"])
        if (right_week == 0):
            vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no such day in2"])
for i in req:
    if i[6] == 'big2':
        day = i[2]
        month = i[3]
        year = i[4]
        pairs = i[5]
        also_find = 0
        right_week = 0
        for j in week6:
            if (day == j[0] and month == j[1] and year == int(j[2])):
                cvobodn = 1
                right_week = 1
                for p in i[5]:
                    if (j[6][p - 1][0] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][0] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 105, 2])
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][36] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][36] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 314, 2])
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][43] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][43] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 328, 2])
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][32] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][32] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 307, 2])
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][42] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][42] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 324, 2])
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][33] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][33] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 309, 2])
                    break
                if cvobodn == 0:
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no free room"])
        if (right_week == 0):
            vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no such day in2"])
for i in req:
    if i[6] == 'any':
        day = i[2]
        month = i[3]
        year = i[4]
        pairs = i[5]
        also_find = 0
        right_week = 0
        for j in week:
            if (day == j[0] and month == j[1] and year == int(j[2])):
                right_week = 1
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][0] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][0] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 502, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][1] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][1] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 506, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][2] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][2] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 508, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][3] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][3] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 509, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][4] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][4] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 511, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][5] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][5] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 512, 6])
                    also_find = 1
                    break
                for p in i[5]:
                    if (j[6][p - 1][6] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][6] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 513, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][7] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][7] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 514, 6])
                    also_find = 1
                    break
                cvobodn = 1
                for p in i[5]:
                    if (j[6][p - 1][8] != ''):
                        cvobodn = 0
                if cvobodn == 1:
                    for p in i[5]:
                        j[6][p - 1][8] = 'CCC'
                    vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 515, 6])
                    also_find = 1
                    break
        if also_find == 0:
            for j in week6:
                if (day == j[0] and month == j[1] and year == int(j[2])):
                    right_week = 1
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][0] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][0] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 105, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][32] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][32] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 307, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][33] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][33] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 309, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][36] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][36] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 314, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][37] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][37] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 317, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][39] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][39] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 318, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][41] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][41] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 322, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][42] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][42] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 324, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][43] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][43] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 328, 2])
                        break
                    if cvobodn == 0:
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no free room"])
        if (right_week == 0):
            vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no such day in6"])
for i in req:
    if i[6] == 'any2':
        day = i[2]
        month = i[3]
        year = i[4]
        pairs = i[5]
        also_find = 0
        right_week = 0
        if also_find == 0:
            for j in week6:
                if (day == j[0] and month == j[1] and year == int(j[2])):
                    right_week = 1
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][0] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][0] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 105, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][32] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][32] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 307, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][33] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][33] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 309, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][36] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][36] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 314, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][37] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][37] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 317, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][39] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][39] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 318, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][41] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][41] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 322, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][42] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][42] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 324, 2])
                        break
                    cvobodn = 1
                    for p in i[5]:
                        if (j[6][p - 1][43] != ''):
                            cvobodn = 0
                    if cvobodn == 1:
                        for p in i[5]:
                            j[6][p - 1][43] = 'CCC'
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8], 328, 2])
                        break
                    if cvobodn == 0:
                        vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no free room"])
        if (right_week == 0):
            vihod.append([i[0] + " " + i[1], int(day), month, year, i[7], i[8]," ", "no such day in6"])
vihod.sort(key = lambda x: x[1])
for i in vihod:
    print(i)
with open('outp.txt', 'w', encoding='utf-8') as f:
    for i in range(len(vihod) - 1):
        for j in range(1, len(vihod[i]) - 1):
            f.write(str(vihod[i][j]) + " ")
        f.write(str(vihod[i][j + 1]) + "\n")
    for j in range(1, len(vihod[len(vihod) - 1]) - 1):
        f.write(str(vihod[len(vihod) - 1][j]) + " ")
    f.write(str(vihod[len(vihod) - 1][j + 1]))
with open('Message.txt', 'w', encoding='utf-8') as f:
    for i in vihod:
        for j in range(0, len(i) - 1):
            f.write(str(i[j]) + " ")
        f.write(str(i[j + 1]) + "\n")
def get_aud():
    korp = ''
    with open("outp.txt", encoding='utf-8') as f:
        inp = f.read()
    inp = inp.split("\n")
    for i in range(0, len(inp)):
        inp[i] = inp[i].split(" ")
    s = set()
    for i in inp:
        s.add(int(i[6]))
    s = list(s)
    s.sort();
    for i in range(len(s)):
        s[i] = str(s[i])
    if len(s) > 1:
        for i in range(0, len(s) - 1):
            korp += s[i] + ", "
        korp += s[len(s) - 1]
    else:
        korp += s[len(s) - 1]

    return (korp, inp)


def nedeed_text1(pdf):
    pdf.cell(200, 13, txt="", ln=1, align="R")
    pdf.cell(190, 6, txt="Начальнику управления безопасности", ln=1, align="R")
    pdf.cell(190, 6, txt="ННГУ им. Н.И. Лобачевского", ln=1, align="R")
    pdf.cell(190, 6, txt="Волкову А.Е.", ln=1, align="R")
    pdf.cell(190, 6, txt="от заместителя директора по учебно-воспитательной работе", ln=1, align="R")
    pdf.cell(190, 6, txt="Института Информационных", ln=1, align="R")
    pdf.cell(190, 6, txt="Технологий, Математики и Механики", ln=1, align="R")
    pdf.cell(190, 6, txt="Кротова Н.В.", ln=1, align="R")
    pdf.cell(190, 8, txt="", ln=1, align="C")
    pdf.cell(190, 15, txt="Служебная записка.", ln=1, align="C")
    pdf.cell(10)
    pdf.multi_cell(180, 6, txt="            Институт Информационных технологий математики"
                               " и механики просит\n "
                               "Вас предоставить доступ в "
                               "аудитории корпуса № " + korp + " ННГУ им. Н.И. Лобачевского "
                                                               "для проведения собрания Студенческого Совета ИИТММ.",
                   align="L")
    pdf.cell(190, 6, txt="", ln=1, align="l")
    return pdf


def nedeed_text2(pdf):
    pdf.cell(10)
    pdf.cell(190, 6, txt="Ответственные:", ln=1, align="l")
    pdf.cell(10)
    pdf.cell(190, 6, txt="Золотых Н.Ю., Борисов Н.А., Кротов Н.В.", ln=1, align="l")
    pdf.cell(190, 6, txt="", ln=1, align="l")
    pdf.cell(10)
    pdf.cell(100, 6, txt="Заместитель директора ИИТММ", ln=1, align="l")
    pdf.cell(10)
    pdf.cell(100, 6, txt="по учебно-воспитательной работе", align="l")
    pdf.cell(40)
    pdf.cell(90, 6, txt="Кротов Н.В.", ln=1, align="r")
    pdf.cell(190, 6, txt="", ln=1, align="l")
    pdf.cell(190, 6, txt="Председатель СС ИИТММ: Саратова Марина +79877471463", ln=1, align="R")
    return pdf

def cycle_out_rooms(pdf, arr, start, end):
    for i in range(start, end):
        pdf.cell(10)
        pdf.cell(190, 6, txt=arr[i][0] + " " + arr[i][1] + " " + arr[i][2] + u" c " + arr[i][3] + u" до " + arr[i][4]
                             + u" ауд. " + arr[i][5] + u" корпуса № " + arr[i][6], ln=1, align="l")
    pdf.cell(190, 6 * (17 - end + start), txt="", ln=1, align="C")
    return pdf

arr = list()
try :
    korp, arr = get_aud()
except Exception:
    print("no such room")
pdf = fpdf.FPDF(format='letter')  # pdf format
pdf.add_page()  # create new page
pdf.add_font('Times', '', 'times.ttf', uni=True)
pdf.set_font("Times", size=14)
pdf = nedeed_text1(pdf)
if len(arr) <= 15:
    pdf = cycle_out_rooms(pdf, arr, 0, len(arr))
    pdf = nedeed_text2(pdf)
else:
    count_pages = len(arr) // 14
    if (len(arr) % 14 != 0):
        count_pages += 1
    pdf = cycle_out_rooms(pdf, arr, 0, 15)
    pdf = nedeed_text2(pdf)
    for i in range(1, count_pages - 1):
        pdf.add_page()
        pdf = nedeed_text1(pdf)
        pdf = cycle_out_rooms(pdf, arr, i * 15, i * 15 + 15)
        pdf = nedeed_text2(pdf)
    pdf.add_page()
    pdf = nedeed_text1(pdf)
    pdf = cycle_out_rooms(pdf, arr, (count_pages - 1) * 15, len(arr))
    pdf = nedeed_text2(pdf)

pdf.output("test.pdf")
